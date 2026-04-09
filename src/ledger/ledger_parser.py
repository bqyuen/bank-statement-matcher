"""
三栏账 Excel 解析器

解析会计三栏账格式的 Excel 文件
样本格式：银行7375辅助三栏账_2月.xlsx
工作表名：辅助三栏账
"""
from typing import List, Optional
from dataclasses import dataclass, field
from datetime import datetime
import openpyxl


@dataclass
class LedgerEntry:
    """三栏账条目"""
    file: str                          # 原始文件路径
    date: Optional[datetime]          # 日期
    voucher_no: str                   # 凭证编号（如"记0004"）
    summary: str                       # 摘要
    amount: float                      # 交易金额（正数）
    direction: str                     # 'income'（收入=借方）/ 'expense'（支出=贷方）
    balance: float                     # 当前余额
    mate: bool = False                 # 是否已匹配
    matched_serial: Optional[str] = None # 匹配的银行流水号
    match_level: Optional[str] = None  # 匹配等级


class LedgerParser:
    """三栏账解析器"""

    def parse(self, file_path: str) -> List[LedgerEntry]:
        """
        解析三栏账Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            LedgerEntry列表
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # 找到工作表（名为"辅助三栏账"）
        if ws.title != '辅助三栏账':
            # 尝试查找
            for sheet_name in wb.sheetnames:
                if '三栏账' in sheet_name or '辅助' in sheet_name:
                    ws = wb[sheet_name]
                    break

        entries = []
        year = None
        month = None
        account_name = ''

        # 扫描行来提取表头信息
        for row_idx in range(1, min(20, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx]]
            row_str = str(row)

            # Row2: "2026-02～2026-02" → 提取年份和月份
            if '～' in row_str or '~' in row_str:
                for cell_val in row:
                    if cell_val and ('～' in str(cell_val) or '~' in str(cell_val)):
                        period = str(cell_val).replace('～', '~').split('~')[0].strip()
                        # period格式: "2026-02"
                        parts = period.split('-')
                        if len(parts) == 2:
                            try:
                                year = int(parts[0])
                                month = int(parts[1])
                            except:
                                pass

            # Row3: 账户名称
            if row_idx == 3:
                for cell_val in row:
                    if cell_val and len(str(cell_val).strip()) > 5:
                        account_name = str(cell_val).strip()
                        break

            # Row4: 列标题（'2026', None, '凭证编号', '摘要', '借方', '贷方', '方向', '金额'）
            if row_idx == 4:
                # 确定列索引
                # A=0(月), B=1(日), C=2(凭证编号), D=3(摘要), E=4(借方), F=5(贷方), G=6(方向), H=7(金额)
                col_month = 0
                col_day = 1
                col_voucher = 2
                col_summary = 3
                col_debit = 4
                col_credit = 5
                col_direction = 6
                col_balance = 7

                # 动态检测列位置（支持列位置不固定的格式）
                def _find_col(row_data, *keywords):
                    for j, v in enumerate(row_data):
                        if v is not None:
                            v_str = str(v).strip()
                            if any(kw in v_str for kw in keywords):
                                return j
                    return -1

                # 在Row4表头行定位各列
                header_row = [cell.value for cell in ws[4]]
                col_month = 0
                col_day = 1
                col_voucher = _find_col(header_row, '凭证编号')
                col_summary = _find_col(header_row, '摘要')
                col_debit = _find_col(header_row, '借方')
                col_credit = _find_col(header_row, '贷方')
                col_direction = _find_col(header_row, '方向')
                col_balance = _find_col(header_row, '金额')

                # 如果动态检测失败（旧格式兼容），使用固定位置
                if col_summary < 0: col_summary = 3
                if col_debit < 0: col_debit = 4
                if col_credit < 0: col_credit = 5
                if col_direction < 0: col_direction = 6
                if col_balance < 0: col_balance = 7

                # 从Row6开始读取数据
                for data_row_idx in range(6, ws.max_row + 1):
                    row_data = [cell.value for cell in ws[data_row_idx]]

                    # 跳过空行
                    if all(v is None or str(v).strip() == '' for v in row_data):
                        continue

                    # 提取各字段
                    m_val = row_data[col_month] if len(row_data) > col_month else None
                    d_val = row_data[col_day] if len(row_data) > col_day else None
                    voucher = row_data[col_voucher] if col_voucher >= 0 and len(row_data) > col_voucher else None
                    summary = row_data[col_summary] if col_summary >= 0 and len(row_data) > col_summary else None
                    debit_val = row_data[col_debit] if col_debit >= 0 and len(row_data) > col_debit else None
                    credit_val = row_data[col_credit] if col_credit >= 0 and len(row_data) > col_credit else None
                    balance_val = row_data[col_balance] if col_balance >= 0 and len(row_data) > col_balance else None

                    # 跳过表头重复行
                    if voucher and str(voucher) in ['凭证编号', '月', '']:
                        continue

                    # 解析日期
                    entry_date = None
                    if m_val is not None and d_val is not None and year is not None:
                        try:
                            m_int = int(str(m_val).strip())
                            d_int = int(str(d_val).strip())
                            entry_date = datetime(year, m_int, d_int)
                        except:
                            pass

                    # 判断收支方向和金额（优先用方向列，其次用借贷金额）
                    amount = 0.0
                    direction = None

                    debit_num = 0.0
                    credit_num = 0.0

                    if debit_val is not None:
                        try:
                            debit_num = float(str(debit_val).replace(',', ''))
                        except:
                            debit_num = 0.0

                    if credit_val is not None:
                        try:
                            credit_num = float(str(credit_val).replace(',', ''))
                        except:
                            credit_num = 0.0

                    if debit_num > 0:
                        amount = debit_num
                        direction = 'income'  # 借方 = 收入
                    elif credit_num > 0:
                        amount = credit_num
                        direction = 'expense'  # 贷方 = 支出
                    else:
                        continue  # 跳过金额为0的行

                    # 解析余额
                    balance = 0.0
                    if balance_val is not None:
                        try:
                            balance = float(str(balance_val).replace(',', ''))
                        except:
                            balance = 0.0

                    # 清洗摘要
                    summary_str = str(summary).replace('\n', ' ').replace('\r', ' ').strip() if summary else ''

                    entry = LedgerEntry(
                        file=file_path,
                        date=entry_date,
                        voucher_no=str(voucher).strip() if voucher else '',
                        summary=summary_str,
                        amount=amount,
                        direction=direction,
                        balance=balance,
                    )
                    entries.append(entry)

        wb.close()
        return entries


def parse_ledger(file_path: str) -> List[LedgerEntry]:
    """快捷函数：解析单个三栏账文件"""
    parser = LedgerParser()
    return parser.parse(file_path)


def parse_ledgers(file_paths: List[str]) -> List[LedgerEntry]:
    """快捷函数：解析多个三栏账文件"""
    all_entries = []
    for fp in file_paths:
        all_entries.extend(parse_ledger(fp))
    return sorted(all_entries, key=lambda e: e.date or datetime.min)
