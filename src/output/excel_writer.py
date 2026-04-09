"""
Excel 核对报告生成器

生成5个Sheet的核对结果Excel报告：
1. 关联明细 - 匹配成功的记录
2. 银行多出 - 银行有、三栏账无
3. 三栏账多出 - 三栏账有、银行无
4. 核对汇总 - 统计汇总
5. 银行存款余额调节表 - 余额调节
"""
from typing import List, Tuple, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _make_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style():
    return {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': _make_border(),
    }


class ExcelReportWriter:
    """
    Excel报告生成器

    Args:
        output_path: 输出Excel文件路径
    """

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = openpyxl.Workbook()

    def write_report(
        self,
        bank_txns: List,
        ledger_entries: List,
        matched: List[Tuple],
        bank_only: List,
        ledger_only: List,
        bank_balance: float,
        ledger_balance: float,
        company_name: str = "",
        account_no: str = "",
        statement_date: str = "",
    ):
        """
        生成完整核对报告

        Args:
            bank_txns: 银行流水列表
            ledger_entries: 三栏账列表
            matched: 匹配结果列表 [(bank_txn, ledger_entry, level), ...]
            bank_only: 银行多出列表
            ledger_only: 三栏账多出列表
            bank_balance: 银行对账单余额
            ledger_balance: 企业日记账余额
            company_name: 编制单位
            account_no: 账户号
            statement_date: 截止日期
        """
        # 删除默认sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        self._write_matched(matched)
        self._write_bank_only(bank_only)
        self._write_ledger_only(ledger_only)
        self._write_summary(
            bank_txns, ledger_entries, matched, bank_only, ledger_only
        )
        self._write_reconciliation(
            bank_txns, ledger_entries, bank_balance, ledger_balance,
            company_name, account_no, statement_date
        )

        self.wb.save(self.output_path)

    # ─── Sheet 1: 关联明细 ───────────────────────────────────────

    def _write_matched(self, matched: List[Tuple]):
        ws = self.wb.create_sheet("关联明细")
        headers = ["银行日期", "凭证编号", "摘要", "对方户名", "银行金额",
                   "收支方向", "匹配等级", "备注"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            for k, v in _header_style().items():
                setattr(cell, k, v)
            cell.font = Font(bold=True, color="FFFFFF")

        for row_idx, (bank_tx, ledger_en, level) in enumerate(matched, 2):
            row_data = [
                bank_tx.date.strftime('%Y-%m-%d') if bank_tx.date else '',
                ledger_en.voucher_no if ledger_en else '',
                ledger_en.summary if ledger_en else bank_tx.summary,
                bank_tx.counterparty if bank_tx else '',
                bank_tx.amount if bank_tx else 0,
                bank_tx.direction if bank_tx else '',
                level,
                f"匹配成功",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 5:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = _make_border()

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 18

    # ─── Sheet 2: 银行多出 ───────────────────────────────────────

    def _write_bank_only(self, bank_only: List):
        ws = self.wb.create_sheet("银行多出")
        headers = ["银行日期", "交易流水号", "摘要", "对方户名", "银行金额", "收支方向", "未匹配原因"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _make_border()

        for row_idx, bank_tx in enumerate(bank_only, 2):
            row_data = [
                bank_tx.date.strftime('%Y-%m-%d') if bank_tx.date else '',
                bank_tx.serial_no if bank_tx else '',
                bank_tx.summary if bank_tx else '',
                bank_tx.counterparty if bank_tx else '',
                bank_tx.amount if bank_tx else 0,
                bank_tx.direction if bank_tx else '',
                "三栏账无对应记录",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 5:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = _make_border()

        for col, w in zip(['A','B','C','D','E','F','G'], [14,22,40,22,14,10,20]):
            ws.column_dimensions[col].width = w

    # ─── Sheet 3: 三栏账多出 ─────────────────────────────────────

    def _write_ledger_only(self, ledger_only: List):
        ws = self.wb.create_sheet("三栏账多出")
        headers = ["日期", "凭证编号", "摘要", "借方", "贷方", "收支方向", "未匹配原因"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7B2C8E", end_color="7B2C8E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _make_border()

        for row_idx, ledger_en in enumerate(ledger_only, 2):
            row_data = [
                ledger_en.date.strftime('%Y-%m-%d') if ledger_en.date else '',
                ledger_en.voucher_no if ledger_en else '',
                ledger_en.summary if ledger_en else '',
                ledger_en.amount if ledger_en.direction == 'income' else 0,
                ledger_en.amount if ledger_en.direction == 'expense' else 0,
                ledger_en.direction if ledger_en else '',
                "银行无对应记录",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx in [4, 5]:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = _make_border()

        for col, w in zip(['A','B','C','D','E','F','G'], [12,12,50,16,16,10,18]):
            ws.column_dimensions[col].width = w

    # ─── Sheet 4: 核对汇总 ────────────────────────────────────────

    def _write_summary(
        self, bank_txns, ledger_entries, matched, bank_only, ledger_only
    ):
        ws = self.wb.create_sheet("核对汇总")
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22

        # 计算统计数据
        bank_income = sum(t.amount for t in bank_txns if t.direction == 'income')
        bank_expense = sum(t.amount for t in bank_txns if t.direction == 'expense')
        ledger_income = sum(e.amount for e in ledger_entries if e.direction == 'income')
        ledger_expense = sum(e.amount for e in ledger_entries if e.direction == 'expense')
        income_diff = ledger_income - bank_income
        expense_diff = ledger_expense - bank_expense
        match_rate = len(matched) / len(bank_txns) * 100 if bank_txns else 0

        summary_data = [
            ("核对项目", "数量/内容", "金额"),
            ("银行流水总记录数", len(bank_txns), "—"),
            ("三栏账总记录数", len(ledger_entries), "—"),
            ("匹配成功记录数", len(matched), "—"),
            ("银行多出记录数", len(bank_only), "—"),
            ("三栏账多出记录数", len(ledger_only), "—"),
            ("匹配率", f"{match_rate:.1f}%", ""),
            ("银行收入合计", f"{sum(1 for t in bank_txns if t.direction == 'income')}笔", f"{bank_income:,.2f}"),
            ("银行支出合计", f"{sum(1 for t in bank_txns if t.direction == 'expense')}笔", f"{bank_expense:,.2f}"),
            ("三栏账收入合计", f"{sum(1 for e in ledger_entries if e.direction == 'income')}笔", f"{ledger_income:,.2f}"),
            ("三栏账支出合计", f"{sum(1 for e in ledger_entries if e.direction == 'expense')}笔", f"{ledger_expense:,.2f}"),
            ("收入差异（企业已收银行未收）", "—", f"{income_diff:,.2f}"),
            ("支出差异（企业已付银行未付）", "—", f"{expense_diff:,.2f}"),
        ]

        for row_idx, (col_a, col_b, col_c) in enumerate(summary_data, 1):
            cell_a = ws.cell(row=row_idx, column=1, value=col_a)
            cell_b = ws.cell(row=row_idx, column=2, value=col_b)
            cell_c = ws.cell(row=row_idx, column=3, value=col_c)

            if row_idx == 1:
                for c in [cell_a, cell_b, cell_c]:
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
                    c.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx == 7:  # 匹配率
                for c in [cell_a, cell_b, cell_c]:
                    c.font = Font(bold=True, size=12)
                    c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif row_idx in [2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13]:
                for c in [cell_a, cell_b, cell_c]:
                    c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            for c in [cell_a, cell_b, cell_c]:
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = _make_border()

    # ─── Sheet 5: 银行存款余额调节表 ─────────────────────────────

    def _write_reconciliation(
        self, bank_txns, ledger_entries, bank_balance, ledger_balance,
        company_name, account_no, statement_date
    ):
        ws = self.wb.create_sheet("银行存款余额调节表")

        # 标题
        ws.merge_cells('A1:D1')
        cell = ws['A1']
        cell.value = "银行存款余额调节表"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # 基本信息
        info = [
            ("编制单位：", company_name or ""),
            ("账户：", account_no or ""),
            ("截止日期：", statement_date or ""),
            ("币种：", "人民币"),
        ]
        for i, (label, val) in enumerate(info, 2):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=val)
            ws.merge_cells(f'B{i}:D{i}')

        for col, w in zip(['A','B','C','D'], [22, 30, 18, 18]):
            ws.column_dimensions[col].width = w

        ws.row_dimensions[6].height = 8

        # 表头
        for col, h in enumerate(["", "项目", "方向", "金额"], 1):
            cell = ws.cell(row=7, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[7].height = 20

        # 计算
        bank_income = sum(t.amount for t in bank_txns if t.direction == 'income')
        bank_expense = sum(t.amount for t in bank_txns if t.direction == 'expense')
        ledger_income = sum(e.amount for e in ledger_entries if e.direction == 'income')
        ledger_expense = sum(e.amount for e in ledger_entries if e.direction == 'expense')

        income_diff = ledger_income - bank_income
        expense_diff = ledger_expense - bank_expense
        adjusted_bank = bank_balance + income_diff - expense_diff
        adjusted_ledger = ledger_balance

        # 银行部分
        bank_section = [
            ("一", "银行对账单余额", "", bank_balance),
            ("", "加：企业已收、银行未收（在途资金）", "收入", income_diff),
            ("", "减：企业已付、银行未付（未达账项）", "支出", expense_diff),
            ("", "", "", None),
            ("", "调整后银行余额", "", adjusted_bank),
        ]
        for row_offset, (num, item, direction, amount) in enumerate(bank_section, 8):
            r = row_offset
            ws.cell(row=r, column=1, value=num).font = Font(bold=True)
            ws.cell(row=r, column=2, value=item)
            ws.cell(row=r, column=3, value=direction)
            c4 = ws.cell(row=r, column=4)
            if amount is not None:
                c4.value = amount
                c4.number_format = '#,##0.00'
                c4.alignment = Alignment(horizontal="right")
            if num == "调整后银行余额":
                for c in range(1, 5):
                    ws.cell(row=r, column=c).font = Font(bold=True)
                    ws.cell(row=r, column=c).fill = PatternFill(
                        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        ws.row_dimensions[13].height = 8

        # 企业部分
        ledger_section = [
            ("二", "企业银行存款日记账余额", "", ledger_balance),
            ("", "加：银行已收、企业未收", "收入", 0.0),
            ("", "减：银行已付、企业未付", "支出", 0.0),
            ("", "", "", None),
            ("", "调整后企业余额", "", adjusted_ledger),
        ]
        for row_offset, (num, item, direction, amount) in enumerate(ledger_section, 14):
            r = row_offset
            ws.cell(row=r, column=1, value=num).font = Font(bold=True)
            ws.cell(row=r, column=2, value=item)
            ws.cell(row=r, column=3, value=direction)
            c4 = ws.cell(row=r, column=4)
            if amount is not None:
                c4.value = amount
                c4.number_format = '#,##0.00'
                c4.alignment = Alignment(horizontal="right")
            if num == "调整后企业余额":
                for c in range(1, 5):
                    ws.cell(row=r, column=c).font = Font(bold=True)
                    ws.cell(row=r, column=c).fill = PatternFill(
                        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        ws.row_dimensions[19].height = 8

        # 平衡结论
        balanced = abs(adjusted_bank - adjusted_ledger) < 0.01
        ws.merge_cells('A20:D20')
        cell = ws['A20']
        if balanced:
            cell.value = (f"✅ 账账相符：调整后银行余额 = 调整后企业余额 = ¥{adjusted_bank:,.2f}")
            cell.font = Font(bold=True, size=11, color="375623")
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            cell.value = (f"🔴 账账不符：银行差 ¥{adjusted_bank:,.2f}，企业差 ¥{adjusted_ledger:,.2f}")
            cell.font = Font(bold=True, size=11, color="C00000")
            cell.fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[20].height = 22

        ws.row_dimensions[22].height = 8
        ws.cell(row=23, column=1, value="核对人：________")
        ws.cell(row=23, column=2, value="复核人：________")
        ws.cell(row=23, column=3, value="日期：________")

        for row in ws.iter_rows(min_row=7, max_row=20, min_col=1, max_col=4):
            for cell in row:
                cell.border = _make_border()
